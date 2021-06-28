import numpy as np
import pandas as pd
import re
from openpyxl import Workbook
import pdfplumber

def gettable2xlsx(pdf, xlsx, num):
    wb = Workbook()
    sheet = wb.active
    with pdfplumber.open(pdf) as pdf:
        for i in range(num):
            page = pdf.pages[i]
            table = page.extract_table()
            for row in table:
                sheet.append(row)
    wb.save(xlsx)

def readcvs2list(name,l):
    df = pd.read_csv(name,usecols=[l],names=None)
    df_li = df.values.tolist()
    result = []
    for s_li in df_li:
        result.append(s_li[0])
    return result

def readexcel2list(name,l):
    df = pd.read_excel(name,usecols=[l],names=None,dtype=str)
    df_li = df.values.tolist()
    result = []
    for s_li in df_li:
        result.append(s_li[0])
    return result

def re4list(list, string):
    pattern = re.compile(string)
    l = []
    for i in list:
        l.append(pattern.search(i)[0])
    return l

def data2xlsx(xlsx, list):
    wb = Workbook()
    sheet = wb.active
    sheet.append(list)
    wb.save(xlsx)

if __name__ == '__main__':
    # gettable2xlsx("stem-list.pdf", "stem.xlsx", 13)
    stem = readexcel2list("stem.xlsx", 1)
    print("STEM:","total:",len(stem),stem)

    CIPcode = readcvs2list("CIPCode2010.csv",1)
    CIPcode = re4list(CIPcode, r'[0-9.]+')
    print("CIPcode","total:",len(CIPcode),CIPcode)

    title = readcvs2list("CIPCode2010.csv", 4)
    definition = readcvs2list("CIPCode2010.csv", 5)

    wb = Workbook()
    sheet = wb.active
    for i in range(len(CIPcode)):
        if (CIPcode[i] in stem):
            continue
        else:
            sheet.append([CIPcode[i],title[i],definition[i]])
    wb.save("nonstem.xlsx")


